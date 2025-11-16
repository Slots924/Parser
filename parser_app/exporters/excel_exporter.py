"""Excel exporter implementation."""
from __future__ import annotations

from pathlib import Path
from typing import Iterable, Optional

try:  # pragma: no cover - handled in runtime branch
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ModuleNotFoundError as exc:  # pragma: no cover - dependency guard
    Workbook = None  # type: ignore[assignment]
    _OPENPYXL_IMPORT_ERROR: Optional[Exception] = exc
else:  # pragma: no cover - exercised when dependency exists
    _OPENPYXL_IMPORT_ERROR = None

from parser_app.config import AppConfig
from parser_app.exporters.base import BaseExporter
from parser_app.models import ProfileRecord


class ExcelExporter(BaseExporter):
    """Export records to an Excel file using :mod:`openpyxl`."""

    def __init__(self, config: AppConfig) -> None:
        if Workbook is None:  # pragma: no cover - requires missing dependency
            raise ImportError(
                "openpyxl is required for Excel export. Install it with 'pip install openpyxl'."
            ) from _OPENPYXL_IMPORT_ERROR
        self.config = config

    def export(self, records: Iterable[ProfileRecord]) -> Path:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = self.config.workbook_title

        headers = list(self.config.header_iter())
        worksheet.append(headers)

        for record in records:
            worksheet.append(record.to_row(headers))

        self._apply_formatting(worksheet, headers)

        output_path = self.config.build_output_path()
        workbook.save(output_path)
        return output_path

    def _apply_formatting(self, worksheet, headers: list[str]) -> None:
        """Apply column widths, fills, and text styles to the worksheet."""

        column_widths = {
            "A": 6.86,
            "B": 37.29,
            "C": 27.00,
            "D": 9.86,
            "E": 11.71,
            "F": 11.00,
            "G": 8.71,
            "H": 21.14,
            "I": 11.86,
            "J": 11.00,
            "K": 8.71,
            "L": 10.43,
            "M": 8.71,
            "N": 4.86,
            "O": 14.00,
            "P": 12.71,
            "Q": 10.86,
            "R": 19.43,
            "S": 8.71,
        }
        for column_letter, width in column_widths.items():
            worksheet.column_dimensions[column_letter].width = width

        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center")

        column_fills = {
            "B": PatternFill(fill_type="solid", fgColor="FFC9DAF8"),
            "C": PatternFill(fill_type="solid", fgColor="FFD0E0E3"),
            "H": PatternFill(fill_type="solid", fgColor="FFC9DAF8"),
            "I": PatternFill(fill_type="solid", fgColor="FFCFE2F3"),
            "R": PatternFill(fill_type="solid", fgColor="FFC9DAF8"),
        }
        clip_columns = {"B", "H", "R"}
        clip_alignment = Alignment(wrap_text=False)

        max_row = worksheet.max_row
        for col_idx in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_idx)
            column_fill = column_fills.get(column_letter)
            should_clip = column_letter in clip_columns

            for row in range(1, max_row + 1):
                cell = worksheet[f"{column_letter}{row}"]
                if row == 1:
                    if should_clip:
                        cell.alignment = Alignment(horizontal="center", wrap_text=False)
                    else:
                        cell.alignment = header_alignment
                    cell.font = header_font
                elif should_clip:
                    cell.alignment = clip_alignment

                if column_fill is not None:
                    cell.fill = column_fill
